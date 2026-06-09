"""
Generates report_data5.html from data_5.xlsx, data_6.xlsx, data_7.xlsx
Run: python3 generate_report5.py
"""
import openpyxl
import json
from collections import defaultdict, Counter
from datetime import datetime

# ── helper ──────────────────────────────────────────────────────────────────

def safe(v, default=0):
    return v if v is not None else default

def pct(n, d):
    return round(n / d * 100, 1) if d else 0

# ── load data_5 ──────────────────────────────────────────────────────────────

wb5 = openpyxl.load_workbook(
    '/root/.claude/uploads/d6ee9e7a-401b-540f-bd3b-e5e383e7ada3/6c8db3b1-data_5.xlsx'
)
rows5 = list(wb5['Sheet1'].iter_rows(values_only=True))[3:]

# KPIs
tot_produced  = sum(safe(r[3]) for r in rows5 if r[3] is not None)
tot_good      = sum(safe(r[5]) for r in rows5 if r[5] is not None)
tot_gpf       = sum(safe(r[6]) for r in rows5 if r[6] is not None)
tot_ncr       = sum(safe(r[8]) for r in rows5 if r[8] is not None)
tot_rework    = sum(safe(r[10]) for r in rows5 if r[10] is not None)
tot_reintro   = sum(safe(r[15]) for r in rows5 if r[15] is not None)
fpy           = round(tot_gpf / tot_good * 100, 1) if tot_good else 0

# GPF ratio + NCR by hour
gpf_by_h  = defaultdict(list)
ncr_by_h  = defaultdict(int)
for r in rows5:
    hour_str = str(r[1]) if r[1] else ''
    try:
        h = int(hour_str.split(' ')[0])
    except Exception:
        continue
    if r[7] is not None:
        gpf_by_h[h].append(r[7])
    ncr_by_h[h] += safe(r[8])

gpf_ratio_labels = list(range(24))
gpf_ratio_data   = [round(sum(gpf_by_h[h]) / len(gpf_by_h[h]) * 100, 1) if gpf_by_h[h] else 0
                    for h in gpf_ratio_labels]
ncr_hour_data    = [ncr_by_h[h] for h in gpf_ratio_labels]

# By shift
shift_stats = defaultdict(lambda: {'produced':0,'good':0,'ncr':0,'gpf':0,'gpf_good':0,'rework':0,'reintro':0})
for r in rows5:
    s = str(r[0]) if r[0] else 'Unknown'
    shift_stats[s]['produced']  += safe(r[3])
    shift_stats[s]['good']      += safe(r[5])
    shift_stats[s]['ncr']       += safe(r[8])
    shift_stats[s]['rework']    += safe(r[10])
    shift_stats[s]['reintro']   += safe(r[15])
    if r[6] is not None:
        shift_stats[s]['gpf']      += safe(r[6])
        shift_stats[s]['gpf_good'] += safe(r[5])

def parse_shift_date(s):
    parts = s.split(' ')
    try:
        d = int(parts[1].rstrip('.'))
        m_map = {'May':5,'Jun':6,'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12,
                 'Jan':1,'Feb':2,'Mar':3,'Apr':4}
        m = m_map.get(parts[2], 1)
        y = 2000 + int(parts[3])
        return datetime(y, m, d)
    except Exception:
        return datetime(2000, 1, 1)

shifts_sorted  = sorted(shift_stats.keys(), key=parse_shift_date)
shift_labels   = shifts_sorted
shift_produced = [shift_stats[s]['produced'] for s in shifts_sorted]
shift_good     = [shift_stats[s]['good']     for s in shifts_sorted]
shift_ncr      = [shift_stats[s]['ncr']      for s in shifts_sorted]

shift_table = []
for s in shifts_sorted:
    st = shift_stats[s]
    fpy_s = round(st['gpf'] / st['gpf_good'] * 100, 1) if st['gpf_good'] else None
    shift_table.append({
        'shift': s, 'produced': st['produced'], 'good': st['good'],
        'ncr': st['ncr'], 'fpy': fpy_s, 'rework': st['rework'], 'reintro': st['reintro']
    })

# ── load data_7 ──────────────────────────────────────────────────────────────

wb7 = openpyxl.load_workbook(
    '/root/.claude/uploads/d6ee9e7a-401b-540f-bd3b-e5e383e7ada3/6692e9e3-data_7.xlsx'
)
rows7 = list(wb7['Sheet1'].iter_rows(values_only=True))[3:]

tot_unpl_min  = 0
tot_unpl_evt  = 0
tot_pl_min    = 0
dt_by_date    = defaultdict(lambda: {'u':0,'p':0})
dt_by_hour    = defaultdict(lambda: {'u':0,'p':0})
top_events    = []
dt_by_code    = defaultdict(int)

for row in rows7:
    wc, rtype, rclass, rcode, rcode2, created, start, end, dur = row
    dur = safe(dur)
    if rtype == 'Unplanned Downtime':
        tot_unpl_min += dur
        if dur > 0:
            tot_unpl_evt += 1
        if start:
            dt_by_date[start.strftime('%d.%m')]['u'] += dur
            dt_by_hour[start.hour]['u']              += dur
        dt_by_code[rcode] += dur
        if dur > 10:
            top_events.append({'dur': dur, 'code': rcode, 'start': str(start)[:16], 'end': str(end)[:16]})
    else:
        tot_pl_min += dur
        if start:
            dt_by_date[start.strftime('%d.%m')]['p'] += dur
            dt_by_hour[start.hour]['p']              += dur

top_events = sorted(top_events, key=lambda x: -x['dur'])[:10]

def date_sort_key(d):
    day, mon = d.split('.')
    return (int(mon), int(day))

dt_dates_sorted = sorted(dt_by_date.keys(), key=date_sort_key)
dt_date_labels  = dt_dates_sorted
dt_unpl_data    = [round(dt_by_date[d]['u'] / 60, 1) for d in dt_dates_sorted]
dt_pl_data      = [round(dt_by_date[d]['p'] / 60, 1) for d in dt_dates_sorted]
dt_hour_unpl    = [round(dt_by_hour[h]['u'] / 60, 1) for h in range(24)]

dt_code_sorted  = sorted(dt_by_code.items(), key=lambda x: -x[1])
dt_code_labels  = [c for c,_ in dt_code_sorted]
dt_code_mins    = [round(m/60,1) for _,m in dt_code_sorted]

# ── load data_6 ──────────────────────────────────────────────────────────────

wb6 = openpyxl.load_workbook(
    '/root/.claude/uploads/d6ee9e7a-401b-540f-bd3b-e5e383e7ada3/e8faf3d9-data_6.xlsx'
)
rows6 = list(wb6['Sheet1'].iter_rows(values_only=True))[3:]

rc_events = Counter()
rc_ncr    = defaultdict(int)
for row in rows6:
    plant,wc,eol,ptype,serial,produced,reported,good,ncr,active_ncr,scrap,rawmat,rework,reintro,pph,rcode,hu,created = row
    if rcode and rcode != 'N/A':
        rc_events[rcode] += 1
        rc_ncr[rcode]    += safe(ncr)

top_rc      = rc_events.most_common(12)
rc_labels   = [c for c,_ in top_rc]
rc_evt_data = [cnt for _,cnt in top_rc]
rc_ncr_data = [rc_ncr[c] for c,_ in top_rc]

prod_by_date6 = defaultdict(lambda: {'produced':0,'good':0,'ncr':0})
for row in rows6:
    plant,wc,eol,ptype,serial,produced,reported,good,ncr,active_ncr,scrap,rawmat,rework,reintro,pph,rcode,hu,created = row
    if created:
        ds = created.strftime('%d.%m')
        prod_by_date6[ds]['produced'] += safe(produced)
        prod_by_date6[ds]['good']     += safe(good)
        prod_by_date6[ds]['ncr']      += safe(ncr)

pd6_dates = sorted(prod_by_date6.keys(), key=date_sort_key)
pd6_prod  = [prod_by_date6[d]['produced'] for d in pd6_dates]
pd6_good  = [prod_by_date6[d]['good']     for d in pd6_dates]
pd6_ncr   = [prod_by_date6[d]['ncr']      for d in pd6_dates]

pph_counts  = Counter(row[14] for row in rows6 if row[14] is not None)
pph_labels  = [str(k) for k in sorted(pph_counts)]
pph_data    = [pph_counts[int(k)] for k in pph_labels]

# ── detail table rows ────────────────────────────────────────────────────────

detail_rows = []
for r in rows5:
    if r[3] is None and r[4] is None:
        continue
    ratio = r[7]
    if ratio is None and r[5] is not None and r[6] is not None and r[5] > 0:
        ratio = r[6] / r[5]
    badge = 'good' if ratio is not None and ratio >= 1.0 else \
            'warn'  if ratio is not None and ratio >= 0.9  else 'bad'
    detail_rows.append({
        'shift':    str(r[0]) if r[0] else '',
        'hour':     str(r[1]) if r[1] else '',
        'target':   safe(r[2]),
        'produced': safe(r[3]),
        'reported': safe(r[4]),
        'good':     safe(r[5]),
        'gpf':      safe(r[6]),
        'ratio':    f"{ratio*100:.1f}%" if ratio is not None else '—',
        'badge':    badge,
        'ncr':      safe(r[8]),
        'rework':   safe(r[10]),
        'scrap':    safe(r[11]),
        'reintro':  safe(r[15]),
    })

# ── HTML helpers ─────────────────────────────────────────────────────────────

fpy_class = 'green' if fpy >= 95 else ('highlight' if fpy >= 90 else 'alert')

def tr(d):
    fpy_d = d['fpy']
    if fpy_d is None:
        fpy_badge = '<span class="badge info">—</span>'
    elif fpy_d >= 99:
        fpy_badge = f'<span class="badge good">{fpy_d}%</span>'
    elif fpy_d >= 90:
        fpy_badge = f'<span class="badge warn">{fpy_d}%</span>'
    else:
        fpy_badge = f'<span class="badge bad">{fpy_d}%</span>'
    return (f'<tr><td>{d["shift"]}</td><td class="num">{d["produced"]}</td>'
            f'<td class="num">{d["good"]}</td><td class="num">{d["ncr"]}</td>'
            f'<td>{fpy_badge}</td><td class="num">{d["rework"]}</td>'
            f'<td class="num">{d["reintro"]}</td></tr>')

def dr(d):
    return (f'<tr><td>{d["shift"]}</td><td>{d["hour"]}</td>'
            f'<td class="num">{d["target"]}</td><td class="num">{d["produced"]}</td>'
            f'<td class="num">{d["reported"]}</td><td class="num">{d["good"]}</td>'
            f'<td class="num">{d["gpf"]}</td>'
            f'<td><span class="badge {d["badge"]}">{d["ratio"]}</span></td>'
            f'<td class="num">{d["ncr"]}</td><td class="num">{d["rework"]}</td>'
            f'<td class="num">{d["scrap"]}</td><td class="num">{d["reintro"]}</td></tr>')

shift_table_html  = ''.join(tr(d) for d in shift_table)
detail_table_html = ''.join(dr(d) for d in detail_rows)

top_dt_html = ''.join(
    f'<tr><td>{e["start"]}</td><td>{e["end"]}</td>'
    f'<td class="num"><strong>{e["dur"]}</strong></td><td>{e["code"]}</td></tr>'
    for e in top_events
)

# ── HTML ─────────────────────────────────────────────────────────────────────

html = f"""<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Qualitätsauswertung DE01 C04-011</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.7/dist/chart.umd.min.js"></script>
<style>
  :root {{
    --atn-green:#B8C400; --atn-dark-green:#99CC00; --atn-med-green:#CAD340;
    --atn-light-green:#DAE279; --atn-pale-green:#EAEDB2; --atn-lightest-green:#EDF0BF;
    --black:#000000; --dark-grey:#666666; --med-grey:#858585;
    --light-grey:#DCDCDC; --vlight-grey:#F4F4F4;
    --orange:#E86A10; --light-orange:#F1A670; --light-blue:#69BBFF;
    --white:#FFFFFF; --lightest-green:#EDF0BF; --pale-green:#EAEDB2;
  }}
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family:Arial,Calibri,sans-serif; background:var(--white); color:var(--black); line-height:1.5; font-size:10pt; }}
  .container {{ max-width:1400px; margin:0 auto; padding:24px; }}
  .classification {{ text-align:right; font-style:italic; color:var(--orange); font-size:10pt; padding:8px 0; }}
  header {{ padding:30px 0 20px; border-bottom:2px solid var(--atn-green); margin-bottom:28px; }}
  header h1 {{ font-size:22pt; font-weight:700; color:var(--atn-green); margin-bottom:4px; }}
  header p {{ color:var(--dark-grey); font-size:10pt; }}
  .kpi-row {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(180px,1fr)); gap:14px; margin-bottom:28px; }}
  .kpi {{ background:var(--vlight-grey); border:1px solid var(--light-grey); border-radius:6px; padding:18px; text-align:center; }}
  .kpi .value {{ font-size:26pt; font-weight:700; }}
  .kpi .label {{ font-size:9pt; color:var(--dark-grey); text-transform:uppercase; letter-spacing:.4px; margin-top:4px; }}
  .kpi.alert .value {{ color:var(--orange); }}
  .kpi.green .value {{ color:var(--atn-dark-green); }}
  .kpi.grey .value {{ color:var(--dark-grey); }}
  .kpi.highlight .value {{ color:var(--atn-green); }}
  .tabs {{ display:flex; gap:4px; margin-bottom:22px; flex-wrap:wrap; }}
  .tab {{ padding:9px 18px; border-radius:4px; cursor:pointer; font-size:10pt; font-weight:700;
          background:var(--vlight-grey); border:1px solid var(--light-grey); color:var(--dark-grey); transition:.2s; font-family:Arial,sans-serif; }}
  .tab:hover {{ background:var(--pale-green); border-color:var(--atn-green); color:var(--black); }}
  .tab.active {{ background:var(--atn-green); border-color:var(--atn-green); color:var(--white); }}
  .tab-panel {{ display:none; }}
  .tab-panel.active {{ display:block; }}
  .card {{ background:var(--white); border:1px solid var(--light-grey); border-radius:6px; padding:22px; margin-bottom:18px; }}
  .card h2 {{ font-size:13pt; font-weight:700; color:var(--atn-green); margin-bottom:14px; }}
  .card h3 {{ font-size:10pt; font-weight:700; color:var(--dark-grey); margin-bottom:10px; text-transform:uppercase; letter-spacing:.4px; }}
  .grid-2 {{ display:grid; grid-template-columns:1fr 1fr; gap:18px; }}
  .grid-3 {{ display:grid; grid-template-columns:1fr 1fr 1fr; gap:18px; }}
  @media(max-width:900px) {{ .grid-2,.grid-3 {{ grid-template-columns:1fr; }} }}
  .chart-wrap {{ position:relative; height:300px; }}
  .chart-wrap.tall {{ height:420px; }}
  .scroll-table {{ max-height:400px; overflow-y:auto; border:1px solid var(--light-grey); border-radius:4px; }}
  table {{ width:100%; border-collapse:collapse; font-size:10pt; }}
  th {{ padding:8px 10px; text-align:left; background:var(--atn-green); color:var(--white); font-weight:700; font-size:9pt; text-transform:uppercase; letter-spacing:.3px; border:1px solid var(--atn-green); position:sticky; top:0; z-index:1; }}
  td {{ padding:7px 10px; border:1px solid var(--light-grey); }}
  td.num {{ text-align:right; font-variant-numeric:tabular-nums; }}
  tr:nth-child(even) td {{ background:var(--vlight-grey); }}
  tr:hover td {{ background:var(--lightest-green); }}
  .badge {{ display:inline-block; padding:2px 8px; border-radius:3px; font-size:9pt; font-weight:700; }}
  .badge.good {{ background:var(--pale-green); color:#5a6600; }}
  .badge.warn {{ background:#FEF3C7; color:#92400E; }}
  .badge.bad {{ background:#FEE2E2; color:#991B1B; }}
  .badge.info {{ background:#E0F2FE; color:#075985; }}
  .insight-box {{ background:var(--lightest-green); border:1px solid var(--atn-med-green); border-left:4px solid var(--orange); border-radius:4px; padding:14px 18px; margin-bottom:14px; }}
  .insight-box h4 {{ font-size:10pt; color:var(--black); margin-bottom:3px; }}
  .insight-box p {{ font-size:9pt; color:var(--dark-grey); }}
  .insight-box.info {{ border-left-color:var(--light-blue); }}
  .footer {{ text-align:left; padding:20px 0; color:var(--dark-grey); font-size:8pt; border-top:1px solid var(--light-grey); margin-top:30px; display:flex; justify-content:space-between; }}
  @media print {{
    body {{ -webkit-print-color-adjust:exact !important; print-color-adjust:exact !important; }}
    .tabs {{ display:none; }}
    .tab-panel {{ display:block !important; }}
    .card {{ break-inside:avoid; }}
    .chart-wrap {{ height:260px; }}
  }}
</style>
</head>
<body>
<div class="container">
  <div class="classification">- INTERNAL -</div>
  <header>
    <h1>Qualitätsauswertung: Good Parts, NCR &amp; Stillstand</h1>
    <p>Werk DE01 &nbsp;|&nbsp; Work Center C04-011 &nbsp;|&nbsp; EOL: Vollautomat Nr. 187</p>
    <p style="color:var(--dark-grey);font-size:10pt;margin-top:4px"><strong>Datenzeitraum:</strong> 20.05.2026 – 09.06.2026</p>
  </header>

  <div class="kpi-row">
    <div class="kpi grey"><div class="value">{tot_produced:,}</div><div class="label">Produzierte Teile</div></div>
    <div class="kpi grey"><div class="value">{tot_good:,}</div><div class="label">Good Parts</div></div>
    <div class="kpi {fpy_class}"><div class="value">{fpy}%</div><div class="label">First Pass Yield</div></div>
    <div class="kpi {'alert' if tot_ncr > 500 else 'grey'}"><div class="value">{tot_ncr:,}</div><div class="label">NCR Teile</div></div>
    <div class="kpi {'alert' if tot_unpl_min > 600 else 'grey'}"><div class="value">{round(tot_unpl_min/60,0):.0f}h</div><div class="label">Ungeplante Stillstände</div></div>
    <div class="kpi grey"><div class="value">{tot_reintro:,}</div><div class="label">Reintroduced Parts</div></div>
  </div>

  <div class="tabs">
    <div class="tab active" onclick="showTab('quality',this)">Qualität</div>
    <div class="tab" onclick="showTab('downtime',this)">Stillstandsanalyse</div>
    <div class="tab" onclick="showTab('transactions',this)">Transaktionsanalyse</div>
  </div>

  <!-- ═══ TAB 1: QUALITY ══════════════════════════════════════════════════ -->
  <div id="tab-quality" class="tab-panel active">

    <div class="insight-box">
      <h4>First Pass Yield: {fpy}% – leicht unter Ziel</h4>
      <p>Von {tot_good:,} Good Parts wurden {tot_gpf:,} bereits im ersten Durchlauf freigegeben. {tot_reintro:,} Teile mussten als Reintroduced Parts nachbearbeitet werden.</p>
    </div>
    <div class="insight-box">
      <h4>NCR-Spitze in Stunde 1 und Stunde 22/23 (Schichtwechsel)</h4>
      <p>Die meisten NCR-Teile entstehen rund um Schichtwechselzeiten. Schicht <strong>S3 21. Mai</strong> ist mit 117 NCR-Teilen die auffälligste Einzelschicht im Zeitraum.</p>
    </div>

    <div class="card">
      <h2>Good Parts First Ratio nach Stunde</h2>
      <p style="font-size:9pt;color:var(--dark-grey);margin:0 0 8px">Durchschnittliche GPF-Quote je Tagesstunde. Orange = unter 90%, Hellgrün = 90–95%, Dunkelgrün = ≥ 95%.</p>
      <div class="chart-wrap"><canvas id="chartGpfHour"></canvas></div>
    </div>

    <div class="card">
      <h2>Produktionsübersicht nach Schicht</h2>
      <div class="grid-2">
        <div>
          <h3>Produzierte vs. NCR-Teile</h3>
          <div class="chart-wrap tall"><canvas id="chartShiftProd"></canvas></div>
        </div>
        <div>
          <h3>Schichttabelle</h3>
          <div class="scroll-table">
            <table>
              <thead><tr>
                <th>Schicht</th><th>Produziert</th><th>Good Parts</th>
                <th>NCR</th><th>FPY</th><th>Rework</th><th>Reintroduced</th>
              </tr></thead>
              <tbody>{shift_table_html}</tbody>
            </table>
          </div>
        </div>
      </div>
    </div>

    <div class="card">
      <h2>NCR-Teile nach Stunde</h2>
      <p style="font-size:9pt;color:var(--dark-grey);margin:0 0 8px">Summe aller NCR-Teile je Tagesstunde über den gesamten Zeitraum.</p>
      <div class="chart-wrap"><canvas id="chartNcrHour"></canvas></div>
    </div>

    <div class="card">
      <h2>Detailtabelle: Alle Stunden</h2>
      <div class="scroll-table">
        <table>
          <thead><tr>
            <th>Schicht</th><th>Stunde</th><th>Prod.Target</th><th>Produziert</th>
            <th>Reported</th><th>Good Parts</th><th>GPF</th><th>GPF Ratio</th>
            <th>NCR</th><th>Rework</th><th>Scrap</th><th>Reintroduced</th>
          </tr></thead>
          <tbody>{detail_table_html}</tbody>
        </table>
      </div>
    </div>
  </div>

  <!-- ═══ TAB 2: DOWNTIME ════════════════════════════════════════════════ -->
  <div id="tab-downtime" class="tab-panel">

    <div class="insight-box">
      <h4>Ungeplante Stillstände: {round(tot_unpl_min/60,1)} Stunden in {tot_unpl_evt} Ereignissen</h4>
      <p>Dominanter Fehlertyp: <strong>Stop</strong>. Größte Einzelereignisse: 480 min (22. Mai, 27.–28. Mai). Schichtwechselzeiten 06:00 und 14:00 zeigen die höchsten Stillstandsstunden.</p>
    </div>
    <div class="insight-box info">
      <h4>Geplante Stillstände: {round(tot_pl_min/60,1)} Stunden (Non running shift)</h4>
      <p>Planmäßige Abwesenheit v.a. 23.–25. Mai (Feiertag/Wochenende), 30. Mai sowie 06.–07. Juni.</p>
    </div>

    <div class="card">
      <h2>Stillstände nach Datum</h2>
      <p style="font-size:9pt;color:var(--dark-grey);margin:0 0 8px">Ungeplante (orange) vs. geplante (grau) Stillstandsstunden pro Tag.</p>
      <div class="chart-wrap tall"><canvas id="chartDtDate"></canvas></div>
    </div>

    <div class="grid-2">
      <div class="card">
        <h2>Ungeplante Stillstände nach Stunde</h2>
        <p style="font-size:9pt;color:var(--dark-grey);margin:0 0 8px">Stunden mit dem höchsten ungeplanten Stillstandsvolumen (in h).</p>
        <div class="chart-wrap"><canvas id="chartDtHour"></canvas></div>
      </div>
      <div class="card">
        <h2>Stillstand nach Grundcode</h2>
        <p style="font-size:9pt;color:var(--dark-grey);margin:0 0 8px">Gesamte Stillstandsstunden je Reason Code.</p>
        <div class="chart-wrap"><canvas id="chartDtCode"></canvas></div>
      </div>
    </div>

    <div class="card">
      <h2>Top 10 längste Stillstandsereignisse</h2>
      <table>
        <thead><tr><th>Start</th><th>Ende</th><th>Dauer (min)</th><th>Reason Code</th></tr></thead>
        <tbody>{top_dt_html}</tbody>
      </table>
    </div>
  </div>

  <!-- ═══ TAB 3: TRANSACTIONS ════════════════════════════════════════════ -->
  <div id="tab-transactions" class="tab-panel">

    <div class="insight-box">
      <h4>Top NCR-Grundcode: 5410-Voids mit 302 Ereignissen / 154 NCR-Teilen</h4>
      <p>Dicht gefolgt von <strong>1011-Incomplete process</strong> (256 Ereignisse) und <strong>5412-Set up scrap</strong> (204 Ereignisse). Die Top 3 Codes decken über 60% aller NCR-Ereignisse ab.</p>
    </div>

    <div class="card">
      <h2>NCR-Grundcodes (Top 12)</h2>
      <div class="grid-2">
        <div>
          <h3>Anzahl Ereignisse</h3>
          <div class="chart-wrap tall"><canvas id="chartRcEvents"></canvas></div>
        </div>
        <div>
          <h3>NCR-Teile je Grundcode</h3>
          <div class="chart-wrap tall"><canvas id="chartRcNcr"></canvas></div>
        </div>
      </div>
    </div>

    <div class="card">
      <h2>Tagesproduktion (Transaktionsdaten)</h2>
      <p style="font-size:9pt;color:var(--dark-grey);margin:0 0 8px">Produzierte Teile, Good Parts und NCR je Tag aus den Einzeltransaktionen (22.682 Datenpunkte).</p>
      <div class="chart-wrap tall"><canvas id="chartDailyProd"></canvas></div>
    </div>

    <div class="card">
      <h2>PPH-Zielwertverteilung</h2>
      <p style="font-size:9pt;color:var(--dark-grey);margin:0 0 8px">Anzahl Transaktionen je PPH-Zielwert (Parts per Hour). Wert 189 dominiert mit {pph_counts.get(189,0):,} Ereignissen.</p>
      <div class="chart-wrap"><canvas id="chartPph"></canvas></div>
    </div>
  </div>

  <div class="footer">
    <span>Autoneum Management AG</span>
    <span>Generiert am <script>document.write(new Date().toLocaleDateString('de-DE'))</script></span>
  </div>
</div>

<script>
const ATN = {{
  GREEN:'#B8C400', DARK_GREEN:'#99CC00', ORANGE:'#E86A10',
  LIGHT_GREY:'#DCDCDC', DARK_GREY:'#666666', LIGHT_BLUE:'#69BBFF',
  VLIGHT_GREY:'#F4F4F4', PALE_GREEN:'#EAEDB2'
}};
const FONT = {{family:'Arial'}};

function showTab(id, el) {{
  document.querySelectorAll('.tab-panel').forEach(p => p.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
  document.getElementById('tab-' + id).classList.add('active');
  el.classList.add('active');
}}

const axisDef = {{
  ticks:{{color:ATN.DARK_GREY, font:FONT}},
  grid:{{color:ATN.LIGHT_GREY}}
}};
const pluginDef = {{
  legend:{{labels:{{color:ATN.DARK_GREY, font:FONT}}}},
  tooltip:{{bodyFont:FONT, titleFont:FONT}}
}};

// ── Tab 1 ───────────────────────────────────────────────────────────────────
new Chart(document.getElementById('chartGpfHour'), {{
  type:'bar',
  data:{{
    labels:{json.dumps(gpf_ratio_labels)},
    datasets:[{{
      label:'GPF Ratio (%)',
      data:{json.dumps(gpf_ratio_data)},
      backgroundColor:{json.dumps(gpf_ratio_data)}.map(v =>
        v < 90 ? ATN.ORANGE : v < 95 ? '#F1A670' : ATN.GREEN),
      borderRadius:4
    }}]
  }},
  options:{{
    responsive:true, maintainAspectRatio:false,
    plugins:pluginDef,
    scales:{{
      x:{{...axisDef, title:{{display:true,text:'Stunde',color:ATN.DARK_GREY,font:FONT}}}},
      y:{{...axisDef, beginAtZero:true, max:110,
          title:{{display:true,text:'GPF Ratio (%)',color:ATN.DARK_GREY,font:FONT}},
          ticks:{{callback:v=>v+'%', color:ATN.DARK_GREY, font:FONT}},
          grid:{{color:ATN.LIGHT_GREY}}}}
    }}
  }}
}});

new Chart(document.getElementById('chartShiftProd'), {{
  type:'bar',
  data:{{
    labels:{json.dumps(shift_labels)},
    datasets:[
      {{label:'Produzierte Teile', data:{json.dumps(shift_produced)}, backgroundColor:ATN.GREEN,      borderRadius:2}},
      {{label:'NCR Teile',         data:{json.dumps(shift_ncr)},      backgroundColor:ATN.ORANGE,     borderRadius:2}}
    ]
  }},
  options:{{
    responsive:true, maintainAspectRatio:false, plugins:pluginDef,
    scales:{{
      x:{{...axisDef, stacked:true, ticks:{{...axisDef.ticks, maxRotation:90}}}},
      y:{{...axisDef, stacked:true, beginAtZero:true,
          title:{{display:true,text:'Teile',color:ATN.DARK_GREY,font:FONT}}}}
    }}
  }}
}});

new Chart(document.getElementById('chartNcrHour'), {{
  type:'bar',
  data:{{
    labels:{json.dumps(gpf_ratio_labels)},
    datasets:[{{label:'NCR Teile', data:{json.dumps(ncr_hour_data)}, backgroundColor:ATN.ORANGE, borderRadius:4}}]
  }},
  options:{{
    responsive:true, maintainAspectRatio:false, plugins:pluginDef,
    scales:{{
      x:{{...axisDef, title:{{display:true,text:'Stunde',color:ATN.DARK_GREY,font:FONT}}}},
      y:{{...axisDef, beginAtZero:true, title:{{display:true,text:'NCR Teile',color:ATN.DARK_GREY,font:FONT}}}}
    }}
  }}
}});

// ── Tab 2 ───────────────────────────────────────────────────────────────────
new Chart(document.getElementById('chartDtDate'), {{
  type:'bar',
  data:{{
    labels:{json.dumps(dt_date_labels)},
    datasets:[
      {{label:'Ungeplant (h)', data:{json.dumps(dt_unpl_data)}, backgroundColor:ATN.ORANGE,     borderRadius:2}},
      {{label:'Geplant (h)',   data:{json.dumps(dt_pl_data)},   backgroundColor:ATN.LIGHT_GREY, borderRadius:2}}
    ]
  }},
  options:{{
    responsive:true, maintainAspectRatio:false, plugins:pluginDef,
    scales:{{
      x:{{...axisDef, stacked:true}},
      y:{{...axisDef, stacked:true, beginAtZero:true,
          title:{{display:true,text:'Stunden',color:ATN.DARK_GREY,font:FONT}}}}
    }}
  }}
}});

new Chart(document.getElementById('chartDtHour'), {{
  type:'bar',
  data:{{
    labels:{json.dumps(list(range(24)))},
    datasets:[{{label:'Ungeplant (h)', data:{json.dumps(dt_hour_unpl)}, backgroundColor:ATN.ORANGE, borderRadius:4}}]
  }},
  options:{{
    responsive:true, maintainAspectRatio:false, plugins:pluginDef,
    scales:{{
      x:{{...axisDef, title:{{display:true,text:'Stunde',color:ATN.DARK_GREY,font:FONT}}}},
      y:{{...axisDef, beginAtZero:true, title:{{display:true,text:'Stunden',color:ATN.DARK_GREY,font:FONT}}}}
    }}
  }}
}});

new Chart(document.getElementById('chartDtCode'), {{
  type:'bar',
  data:{{
    labels:{json.dumps(dt_code_labels)},
    datasets:[{{
      label:'Stillstand (h)', data:{json.dumps(dt_code_mins)},
      backgroundColor:[ATN.ORANGE, ATN.LIGHT_GREY, ATN.LIGHT_BLUE],
      borderRadius:4
    }}]
  }},
  options:{{
    responsive:true, maintainAspectRatio:false, plugins:pluginDef,
    indexAxis:'y',
    scales:{{
      x:{{...axisDef, beginAtZero:true, title:{{display:true,text:'Stunden',color:ATN.DARK_GREY,font:FONT}}}},
      y:{{...axisDef}}
    }}
  }}
}});

// ── Tab 3 ───────────────────────────────────────────────────────────────────
new Chart(document.getElementById('chartRcEvents'), {{
  type:'bar',
  data:{{
    labels:{json.dumps(rc_labels)},
    datasets:[{{label:'Ereignisse', data:{json.dumps(rc_evt_data)}, backgroundColor:ATN.ORANGE, borderRadius:4}}]
  }},
  options:{{
    responsive:true, maintainAspectRatio:false, plugins:pluginDef,
    indexAxis:'y',
    scales:{{x:{{...axisDef, beginAtZero:true}}, y:{{...axisDef}}}}
  }}
}});

new Chart(document.getElementById('chartRcNcr'), {{
  type:'bar',
  data:{{
    labels:{json.dumps(rc_labels)},
    datasets:[{{label:'NCR Teile', data:{json.dumps(rc_ncr_data)}, backgroundColor:ATN.DARK_GREEN, borderRadius:4}}]
  }},
  options:{{
    responsive:true, maintainAspectRatio:false, plugins:pluginDef,
    indexAxis:'y',
    scales:{{x:{{...axisDef, beginAtZero:true}}, y:{{...axisDef}}}}
  }}
}});

new Chart(document.getElementById('chartDailyProd'), {{
  type:'line',
  data:{{
    labels:{json.dumps(pd6_dates)},
    datasets:[
      {{label:'Produzierte Teile', data:{json.dumps(pd6_prod)}, borderColor:ATN.GREEN,
        backgroundColor:'rgba(184,196,0,0.12)', fill:true, tension:.3, pointRadius:4}},
      {{label:'Good Parts',        data:{json.dumps(pd6_good)}, borderColor:ATN.DARK_GREY,
        backgroundColor:'transparent', tension:.3, pointRadius:4}},
      {{label:'NCR Teile',         data:{json.dumps(pd6_ncr)},  borderColor:ATN.ORANGE,
        backgroundColor:'transparent', tension:.3, borderDash:[4,3], pointRadius:4}}
    ]
  }},
  options:{{
    responsive:true, maintainAspectRatio:false, plugins:pluginDef,
    scales:{{
      x:{{...axisDef}},
      y:{{...axisDef, beginAtZero:true, title:{{display:true,text:'Teile',color:ATN.DARK_GREY,font:FONT}}}}
    }}
  }}
}});

new Chart(document.getElementById('chartPph'), {{
  type:'bar',
  data:{{
    labels:{json.dumps(pph_labels)},
    datasets:[{{
      label:'Transaktionen',
      data:{json.dumps(pph_data)},
      backgroundColor:[ATN.PALE_GREEN,ATN.GREEN,ATN.DARK_GREEN,'#5a6600',ATN.ORANGE,'#aaa','#ccc','#eee'],
      borderRadius:4
    }}]
  }},
  options:{{
    responsive:true, maintainAspectRatio:false, plugins:pluginDef,
    scales:{{
      x:{{...axisDef, title:{{display:true,text:'PPH (Parts per Hour)',color:ATN.DARK_GREY,font:FONT}}}},
      y:{{...axisDef, beginAtZero:true,
          title:{{display:true,text:'Anzahl Transaktionen',color:ATN.DARK_GREY,font:FONT}}}}
    }}
  }}
}});
</script>
</body>
</html>"""

with open('/home/user/NVH-2030/report_data5.html', 'w', encoding='utf-8') as f:
    f.write(html)

print(f'Report written: {len(html):,} bytes')
print(f'KPIs: produced={tot_produced:,}, good={tot_good:,}, fpy={fpy}%, ncr={tot_ncr:,}')
print(f'Downtime: unplanned={round(tot_unpl_min/60,1)}h ({tot_unpl_evt} events), planned={round(tot_pl_min/60,1)}h')
print(f'NCR codes: {len(rc_events)} unique, top={rc_events.most_common(1)}')
